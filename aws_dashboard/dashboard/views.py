from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
import boto3
from openpyxl import Workbook
from datetime import datetime, timezone
from .aws_utils import *
from .url_checker import *

# INDEX VIEW
def index(request):
    return render(request, 'dashboard/index.html')

# -----------------------EC2 view--------------------
def ec2_today_status(request):
    ec2 = boto3.client('ec2')
    instances = ec2.describe_instances()

    today = datetime.now(timezone.utc).date()
    today_instances = []
    other_instances = []

    for reservation in instances['Reservations']:
        for instance in reservation['Instances']:
            launch_time = instance['LaunchTime'].date()
            instance_info = {
                'id': instance['InstanceId'],
                'state': instance['State']['Name'],
                'launch_time': instance['LaunchTime'],
                'name': next((tag['Value'] for tag in instance.get('Tags', []) if tag['Key'] == 'Name'), 'N/A')
            }
            if launch_time == today:
                today_instances.append(instance_info)
            else:
                other_instances.append(instance_info)

    return render(request, 'dashboard/partials/ec2.html', {
        'today_instances': today_instances,
        'other_instances': other_instances
    })

# ----------------------Download EC2 Status as Excel---------------------------------
def export_ec2_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "EC2 Status (Today)"
    ec2 = boto3.client('ec2')
    reservations = ec2.describe_instances()['Reservations']
    today = datetime.now(timezone.utc).date()

    ws.append(['Instance ID', 'Name', 'Launch Time', 'Status', 'Launched Today'])
    for res in reservations:
        for inst in res['Instances']:
            inst_id = inst['InstanceId']
            launch_time = inst['LaunchTime']
            name = next((t['Value'] for t in inst.get('Tags', []) if t['Key'] == 'Name'), '')
            status = inst['State']['Name']
            launched_today = 'Yes' if launch_time.date() == today else 'No'
            ws.append([inst_id, name, str(launch_time), status, launched_today])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=ec2_instances.xlsx"
    wb.save(response)
    return response

# -------------------ASG SUMMARY PAGE---------------------
def get_all_auto_scaling_groups():
    client = boto3.client('autoscaling')
    asgs = []
    next_token = None

    while True:
        if next_token:
            response = client.describe_auto_scaling_groups(NextToken=next_token)
        else:
            response = client.describe_auto_scaling_groups()

        asgs.extend(response.get('AutoScalingGroups', []))
        next_token = response.get('NextToken')
        if not next_token:
            break

    return asgs

def asg_summary(request):
    autoscaling = boto3.client('autoscaling')

    # Get all ASGs
    asgs = get_all_auto_scaling_groups()
    asg_count = len(asgs)

    # Prepare ASG list details
    asg_list = []
    for asg in asgs:
        asg_list.append({
            'name': asg.get('AutoScalingGroupName'),
            'min_size': asg.get('MinSize'),
            'max_size': asg.get('MaxSize'),
            'desired_capacity': asg.get('DesiredCapacity'),
            'instance_count': len(asg.get('Instances', []))
        })

    return render(request, 'dashboard/partials/asg.html', {
        'asg_list': asg_list,
        'asg_count': asg_count
    })

# -------------AMI SUMMARY PAGE---------------------
def ami_summary(request):
    ec2 = boto3.client('ec2')
    # AMIs (Owned by you)
    amis = ec2.describe_images(Owners=['self'])
    ami_count = len(amis['Images'])
    ami_list = []
    for ami in amis['Images']:
        ami_list.append({
            'image_id': ami.get('ImageId'),
            'name': ami.get('Name'),
            'creation_date': ami.get('CreationDate'),
            'state': ami.get('State')
        })

    return render(request, 'dashboard/partials/ami.html', {
        'ami_list': ami_list,'ami_count': ami_count
    })

# ------------------SNAPSHOTS SUMMARY PAGE---------------------------
def snapshots_summary(request):
    ec2 = boto3.client('ec2')
    # Snapshots (Owned by you)
    snapshots = ec2.describe_snapshots(OwnerIds=['self'])
    snapshots_count = len(snapshots['Snapshots'])
    snapshot_list = []
    for snap in snapshots['Snapshots']:
        snapshot_list.append({
            'snapshot_id': snap.get('SnapshotId'),
            'start_time': str(snap.get('StartTime')),
            'volume_size': snap.get('VolumeSize'),
            'state': snap.get('State')
        })

    return render(request, 'dashboard/partials/snapshots.html', {
        'snapshot_list': snapshot_list,'snapshots_count': snapshots_count
    })

# LOAD BALANCER SUMMARY PAGE
def lb_summary(request):
    elbv2 = boto3.client('elbv2')
    lbs = elbv2.describe_load_balancers()
    lb_list = []
    for lb in lbs['LoadBalancers']:
        lb_list.append({
            'name': lb.get('LoadBalancerName'),
            'type': lb.get('Type'),
            'state': lb['State']['Code']
        })

    return render(request, 'dashboard/partials/load_balancer.html', {
        'lb_list': lb_list,
    })

# LAMBDA SUMMARY PAGE
def get_all_lambda_functions(lambda_client):
    lambda_functions = []
    next_marker = None

    while True:
        # Properly handle pagination using Marker
        if next_marker:
            response = lambda_client.list_functions(Marker=next_marker)
        else:
            response = lambda_client.list_functions()
        lambda_functions.extend(response.get('Functions', []))
        next_marker = response.get('NextMarker')
        if not next_marker:
            break

    return lambda_functions


def lambda_summary(request):
    lambda_client = boto3.client('lambda')
    lambda_functions = get_all_lambda_functions(lambda_client)
    lambda_list = []
    for fn in lambda_functions:
        lambda_list.append({
            'name': fn.get('FunctionName', 'N/A'),
            'runtime': fn.get('Runtime', 'N/A'),
            'last_modified': fn.get('LastModified', 'N/A'),
            'memory_size': fn.get('MemorySize', 'N/A'),
        })

    return render(request, 'dashboard/partials/lambda.html', {
        'lambda_list': lambda_list,
    })


# Download Lambda Summary as Excel
def export_lambda_excel(request):
    lambda_client = boto3.client('lambda')
    lambda_functions = get_all_lambda_functions(lambda_client)

    wb = Workbook()
    ws = wb.active
    ws.title = "Lambda Functions"

    # Header
    headers = ['Function Name', 'Runtime', 'Handler', 'Memory (MB)', 'Last Modified']
    ws.append(headers)

    # Data rows
    for fn in lambda_functions:
        ws.append([
            fn.get('FunctionName', 'N/A'),
            fn.get('Runtime', 'N/A'),
            fn.get('Handler', 'N/A'),
            fn.get('MemorySize', 'N/A'),
            fn.get('LastModified', 'N/A')  
        ])

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Lambda_Summary.xlsx'
    wb.save(response)
    return response

# Download ASG Summary as Excel
def export_asg_excel(request):
    asgs = get_all_auto_scaling_groups()

    wb = Workbook()
    ws = wb.active
    ws.title = "Auto Scaling Groups"
    ws.append(["Name", "Min Size", "Max Size", "Desired Capacity", "Instance Count"])

    for asg in asgs:
        ws.append([
            asg.get('AutoScalingGroupName'),
            asg.get('MinSize'),
            asg.get('MaxSize'),
            asg.get('DesiredCapacity'),
            len(asg.get('Instances', []))
        ])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=asg_summary.xlsx'
    wb.save(response)
    return response

# Download AMI Summary as Excel
def export_ami_excel(request):
    ec2 = boto3.client('ec2')
    amis = ec2.describe_images(Owners=['self'])

    wb = Workbook()
    ws = wb.active
    ws.title = "AMIs"
    ws.append(["Image ID", "Name", "Creation Date", "State"])

    for ami in amis['Images']:
        ws.append([
            ami.get('ImageId'),
            ami.get('Name'),
            ami.get('CreationDate'),
            ami.get('State')
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"ami_summary_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# Download Snapshot Summary as Excel
def export_snapshot_excel(request):
    ec2 = boto3.client('ec2')
    snapshots = ec2.describe_snapshots(OwnerIds=['self'])

    wb = Workbook()
    ws = wb.active
    ws.title = "Snapshots"
    ws.append(["Snapshot ID", "Start Time", "Size (GB)", "State"])

    for snap in snapshots['Snapshots']:
        ws.append([
            snap.get('SnapshotId'),
            str(snap.get('StartTime')),
            snap.get('VolumeSize'),
            snap.get('State')
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"snapshots_summary_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# Download Load Balancer Summary as Excel
def export_lb_excel(request):
    elbv2 = boto3.client('elbv2')
    lbs = elbv2.describe_load_balancers()
    lb_list = []
    for lb in lbs['LoadBalancers']:
        lb_list.append({
            'name': lb.get('LoadBalancerName'),
            'type': lb.get('Type'),
            'state': lb['State']['Code']
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "Load Balancers"

    # Write headers
    ws.append(["Name", "Type", "State"])

    # Write data
    for lb in lb_list:
        ws.append([
            lb.get('name', ''),
            lb.get('type', ''),
            lb.get('state', '')
        ])

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=load_balancers.xlsx'
    wb.save(response)
    return response

# URL HEALTH CHECK PAGE
def url_health_check(request):
    url_stgsqlup = check_urls_from_file('dashboard/stgsqlupurl.txt')
    url_results = check_urls_from_file('dashboard/dev2url.txt')

    return render(request, 'dashboard/partials/url.html', {
        'url_results': url_results,'url_stgsqlup':url_stgsqlup
    })

# Download URL Health Status as Excel
def export_url_health_excel(request):
    url_stgsqlup = check_urls_from_file('dashboard/stgsqlupurl.txt')
    url_results = check_urls_from_file('dashboard/dev2url.txt')

    wb = Workbook()
    ws = wb.active
    ws.title = "URL Health Status"
    ws.append(["URL", "Status", "Status Code"])

    for result in url_results:
        ws.append([result['url'], result['status'], result['status_code']])
    
    for result in url_stgsqlup:
        ws.append([result['url'], result['status'], result['status_code']])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"url_health_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# EC2 UTILIZATION PAGE
def ec2_utilization(request):
    metrics = get_ec2_utilization()
    data = get_ec2_utilization()
    return JsonResponse({'metrics': metrics,'instances': data})

def ec2_utilization_ui(request):
    return render(request, 'dashboard/partials/ec2_util.html')

# Download EC2 Utilization as Excel
def export_ec2_utilization_excel(request):
    metrics = get_ec2_utilization()  # Call your existing function

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "EC2 Utilization"

    # Header row
    ws.append(["Instance ID", "Name", "CPU Utilization (%)"])

    # Data rows
    for m in metrics:
        ws.append([
            m.get('instance_id', ''),
            m.get('name', ''),
            m.get('cpu', 0.0)
        ])

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=ec2_utilization.xlsx'
    wb.save(response)
    return response

# RDS UTILIZATION PAGE
def rds_utilization(request):
    metrics = get_rds_utilization()
    data = get_rds_utilization()
    return JsonResponse({'metrics': metrics, 'instances': data})

def rds_utilization_ui(request):
    return render(request, 'dashboard/partials/rds_redis.html')

def redis_utilization(request):
    data = get_redis_utilization()
    return JsonResponse({'metrics': data})

def redis_utilization_ui(request):
    return render(request, 'dashboard/partials/rds_redis.html')


# Download RDS Utilization as Excel
def export_rds_utilization_excel(request):
    metrics = get_rds_utilization()
    datas = get_redis_utilization()

    wb = Workbook()
    ws = wb.active
    ws.title = "Rds&Redis Utilization"

    ws.append(['Resources','Name', 'CPU Utilization (%)', 'Memory'])

    for metric in datas:
        ws.append([
            'Redis',
            metric['cluster'],
            metric['cpu'],
            metric['memory']
        ])

    for metric in metrics:
        ws.append([
            'RDS',
            metric['db_id'],
            metric['cpu'],
            metric['memory']
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Rds&Redis_Utilization.xlsx'
    wb.save(response)
    return response



# ECS UTILIZATION PAGE
def ecs_utilization(request):
    metrics = get_ecs_utilization()
    return JsonResponse({'metrics': metrics})

def ecs_utilization_ui(request):
    return render(request, 'dashboard/partials/ecs.html')

# Download ECS Utilization as Excel
def export_ecs_utilization_excel(request):
    metrics = get_ecs_utilization()

    wb = Workbook()
    ws = wb.active
    ws.title = "ECS Utilization"

    # Header row
    ws.append(['Cluster', 'Service', 'CPU Utilization (%)', 'Memory Utilization (%)'])

    # Data rows
    for m in metrics:
        ws.append([
            m['cluster'],
            m['service'],
            m['cpu'],
            m['memory']
        ])

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=ECS_Utilization.xlsx'
    wb.save(response)
    return response

# Target Group summary page

def list_target_groups_with_health_and_lb():
    elbv2 = boto3.client('elbv2')
    paginator = elbv2.get_paginator('describe_target_groups')

    result = []

    for page in paginator.paginate():
        for tg in page['TargetGroups']:
            tg_arn = tg.get('TargetGroupArn')
            tg_name = tg.get('TargetGroupName', 'N/A')

            # Fetch associated Load Balancers
            lb_arns = tg.get('LoadBalancerArns', [])
            if lb_arns:
                lb_names = []
                for lb_arn in lb_arns:
                    lb_info = elbv2.describe_load_balancers(LoadBalancerArns=[lb_arn])
                    for lb in lb_info.get('LoadBalancers', []):
                        lb_names.append(lb.get('LoadBalancerName'))
            else:
                lb_names = ['Not associated']

            # Fetch target health
            try:
                health_response = elbv2.describe_target_health(TargetGroupArn=tg_arn)
                targets = []
                for th in health_response['TargetHealthDescriptions']:
                    targets.append({
                        'id': th['Target']['Id'],
                        'port': th['Target']['Port'],
                        'state': th['TargetHealth']['State'],
                        'reason': th['TargetHealth'].get('Reason', ''),
                        'description': th['TargetHealth'].get('Description', '')
                    })
            except Exception as e:
                targets = [{'id': 'Error', 'port': '-', 'state': str(e), 'reason': '', 'description': ''}]

            result.append({
                'name': tg_name,
                'protocol': tg.get('Protocol', 'N/A'),
                'port': tg.get('Port', 'N/A'),
                'vpc_id': tg.get('VpcId', 'N/A'),
                'target_type': tg.get('TargetType', 'N/A'),
                'load_balancers': lb_names,
                'targets': targets
            })

    return result

def target_group_list(request):
    groups = list_target_groups_with_health_and_lb()
    tg_count = len(groups)
    return render(request, 'dashboard/partials/target_groups.html', {
        'target_groups': groups,
        'tg_count': tg_count
    })


def asg_utilization(request):
    metrics = get_asg_utilization()
    return JsonResponse({'metrics': metrics})

def asg_utilization_ui(request):
    return render(request, 'dashboard/partials/asg_util.html')